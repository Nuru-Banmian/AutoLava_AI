from collections.abc import Iterable
from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell


def _safe_text(value: object) -> object:
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def build_ledger_workbook(records: Iterable[dict], *, include_wash_count: bool = True) -> bytes:
    records = list(records)
    workbook = Workbook(write_only=True)
    summary = workbook.create_sheet(title="经营记录")
    summary_headers = ["日期", "状态", "总收入"]
    if include_wash_count:
        summary_headers.append("洗车")
    summary_headers.extend(["天气", "事件", "记录人", "最后修改人"])
    summary.append(summary_headers)

    detail = workbook.create_sheet(title="收入明细")
    detail.append(["日期", "收入项目", "计入总额", "排序", "金额"])

    def money_cell(sheet, value: int) -> WriteOnlyCell:
        cell = WriteOnlyCell(sheet, value=int(value))
        cell.number_format = "€#,##0"
        return cell

    for record in records:
        summary_row = [
            date.fromisoformat(record["date"]),
            record["is_open"],
            money_cell(summary, record["daily_revenue"]),
        ]
        if include_wash_count:
            summary_row.append(record["wash_count"])
        summary_row.extend(
            [
                _safe_text(record["weather"]),
                _safe_text(record["activity"]),
                _safe_text(record["created_by_name"]),
                _safe_text(record["updated_by_name"]),
            ]
        )
        summary.append(summary_row)
        for item in sorted(record["items"], key=lambda value: (value["sort_order"], value["id"])):
            detail.append(
                [
                    date.fromisoformat(record["date"]),
                    _safe_text(item["category_name"]),
                    item["include_in_total"],
                    item["sort_order"],
                    money_cell(detail, item["amount"]),
                ]
            )
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
