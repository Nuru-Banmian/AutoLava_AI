from datetime import date, datetime
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import select

from app.models.identity import StoreMember, User
from app.models.ledger import DailyIncomeItem, IncomeCategory, StoreDailyRecord


@pytest.mark.parametrize(
    ("method", "path"),
    [
        ("get", "/api/database/1/history"),
        ("post", "/api/database/1/history/1/rollback"),
        ("post", "/api/database/1/rollback/1"),
    ],
)
async def test_history_and_rollback_routes_do_not_exist(
    auth_client, method: str, path: str
) -> None:
    response = await auth_client.request(method, path)
    assert response.status_code == 404


async def test_database_records_route_remains_available_for_assigned_user(
    auth_client, store_factory, db_session
) -> None:
    store = await store_factory(name="Records")
    user = await db_session.scalar(select(User).where(User.username == "authenticated"))
    db_session.add(StoreMember(store_id=store.id, user_id=user.id))
    await db_session.flush()

    response = await auth_client.get(f"/api/database/{store.id}/records")

    assert response.status_code == 200


async def test_database_summary_and_export_use_integer_money(
    auth_client, store_factory, db_session
) -> None:
    store = await store_factory(name="Integer Records")
    user = await db_session.scalar(select(User).where(User.username == "authenticated"))
    assert user is not None
    category = IncomeCategory(
        store_id=store.id,
        name="Cash",
        include_in_total=True,
        is_active=True,
        sort_order=0,
    )
    db_session.add_all([StoreMember(store_id=store.id, user_id=user.id), category])
    await db_session.flush()
    record = StoreDailyRecord(
        store_id=store.id,
        date=date(2026, 7, 18),
        daily_revenue=321,
        income_mode="composed",
        is_open="营业",
        weather_edited=False,
        created_by=user.id,
        updated_by=user.id,
    )
    db_session.add(record)
    await db_session.flush()
    db_session.add(
        DailyIncomeItem(
            record_id=record.id,
            category_id=category.id,
            category_name="Cash",
            include_in_total=True,
            sort_order=0,
            amount=321,
        )
    )
    await db_session.flush()

    page = await auth_client.get(f"/api/database/{store.id}/records")
    exported = await auth_client.get(f"/api/database/{store.id}/export.xlsx")

    assert page.status_code == exported.status_code == 200
    assert page.json()["sum_daily_revenue"] == 321
    assert isinstance(page.json()["sum_daily_revenue"], int)
    workbook = load_workbook(BytesIO(exported.content), read_only=False)
    sheet = workbook["经营记录"]
    assert sheet.cell(row=2, column=3).value == 321
    assert sheet.cell(row=2, column=3).number_format == "€#,##0"
    detail = workbook["收入明细"]
    assert detail.cell(row=2, column=5).value == 321
    assert detail.cell(row=2, column=5).number_format == "€#,##0"


async def test_export_uses_saved_income_item_snapshots_after_current_category_changes(
    auth_client, store_factory, db_session
) -> None:
    store = await store_factory(name="Historical Export")
    user = await db_session.scalar(select(User).where(User.username == "authenticated"))
    assert user is not None
    category = IncomeCategory(
        store_id=store.id,
        name="Original current name",
        include_in_total=True,
        is_active=True,
        sort_order=0,
    )
    db_session.add_all([StoreMember(store_id=store.id, user_id=user.id), category])
    await db_session.flush()
    record = StoreDailyRecord(
        store_id=store.id,
        date=date(2026, 7, 1),
        daily_revenue=150,
        income_mode="composed",
        is_open="营业",
        weather_edited=False,
        created_by=user.id,
        updated_by=user.id,
    )
    db_session.add(record)
    await db_session.flush()
    db_session.add(
        DailyIncomeItem(
            record_id=record.id,
            category_id=category.id,
            category_name="Historical name",
            include_in_total=True,
            sort_order=0,
            amount=150,
        )
    )
    await db_session.flush()

    category.name = "Renamed current category"
    category.include_in_total = False
    category.sort_order = 9
    await db_session.flush()

    exported = await auth_client.get(f"/api/database/{store.id}/export.xlsx")

    assert exported.status_code == 200
    workbook = load_workbook(BytesIO(exported.content), read_only=False)
    summary = workbook["经营记录"]
    assert [cell.value for cell in summary[1]] == [
        "日期",
        "状态",
        "总收入",
        "洗车",
        "天气",
        "事件",
        "记录人",
        "最后修改人",
    ]
    detail = workbook["收入明细"]
    assert [detail.cell(row=2, column=index).value for index in range(1, 6)] == [
        datetime(2026, 7, 1),
        "Historical name",
        True,
        0,
        150,
    ]
    assert detail.cell(row=2, column=5).number_format == "€#,##0"


async def test_database_context_and_export_follow_the_requested_store_wash_setting(
    auth_client, store_factory, db_session
) -> None:
    user = await db_session.scalar(select(User).where(User.username == "authenticated"))
    assert user is not None
    store = await store_factory(name="Wash-disabled records")
    other = await store_factory(name="Wash-enabled records")
    store.wash_count_enabled = False
    other.wash_count_enabled = True
    db_session.add(StoreMember(store_id=store.id, user_id=user.id))
    db_session.add(
        StoreDailyRecord(
            store_id=store.id,
            date=date(2026, 7, 20),
            daily_revenue=240,
            wash_count=8,
            is_open="营业",
            weather_edited=False,
            created_by=user.id,
            updated_by=user.id,
        )
    )
    await db_session.flush()

    disabled_page = await auth_client.get(
        f"/api/database/{store.id}/records?missing_wash_count=true"
    )
    disabled_export = await auth_client.get(
        f"/api/database/{store.id}/export.xlsx?missing_wash_count=true"
    )

    assert disabled_page.status_code == disabled_export.status_code == 200
    assert disabled_page.json()["total"] == 1
    assert "wash_count" not in disabled_page.json()["items"][0]
    disabled_workbook = load_workbook(BytesIO(disabled_export.content), read_only=False)
    assert [cell.value for cell in disabled_workbook["经营记录"][1]] == [
        "日期",
        "状态",
        "总收入",
        "天气",
        "事件",
        "记录人",
        "最后修改人",
    ]

    store.wash_count_enabled = True
    await db_session.flush()
    reenabled_page = await auth_client.get(f"/api/database/{store.id}/records")
    filtered_page = await auth_client.get(
        f"/api/database/{store.id}/records?missing_wash_count=true"
    )
    reenabled_export = await auth_client.get(
        f"/api/database/{store.id}/export.xlsx"
    )

    assert reenabled_page.json()["items"][0]["wash_count"] == 8
    assert filtered_page.json()["total"] == 0
    reenabled_workbook = load_workbook(BytesIO(reenabled_export.content), read_only=False)
    assert [cell.value for cell in reenabled_workbook["经营记录"][1]][:4] == [
        "日期",
        "状态",
        "总收入",
        "洗车",
    ]


async def test_database_records_and_export_filter_by_early_close_status(
    auth_client, store_factory, db_session
) -> None:
    store = await store_factory(name="Early close records")
    user = await db_session.scalar(select(User).where(User.username == "authenticated"))
    assert user is not None
    db_session.add_all(
        [
            StoreMember(store_id=store.id, user_id=user.id),
            StoreDailyRecord(
                store_id=store.id,
                date=date(2026, 7, 21),
                daily_revenue=240,
                wash_count=8,
                is_open="提前休息",
                weather_edited=False,
                created_by=user.id,
                updated_by=user.id,
            ),
        ]
    )
    await db_session.flush()

    page = await auth_client.get(
        f"/api/database/{store.id}/records", params={"status": "提前休息"}
    )
    exported = await auth_client.get(
        f"/api/database/{store.id}/export.xlsx", params={"status": "提前休息"}
    )
    legacy_page = await auth_client.get(
        f"/api/database/{store.id}/records", params={"status": "天气停业"}
    )
    legacy_export = await auth_client.get(
        f"/api/database/{store.id}/export.xlsx", params={"status": "天气停业"}
    )

    assert page.status_code == exported.status_code == 200
    assert page.json()["total"] == 1
    assert page.json()["items"][0]["is_open"] == "提前休息"
    workbook = load_workbook(BytesIO(exported.content), read_only=False)
    assert workbook["经营记录"].cell(row=2, column=2).value == "提前休息"
    assert legacy_page.status_code == legacy_export.status_code == 422
